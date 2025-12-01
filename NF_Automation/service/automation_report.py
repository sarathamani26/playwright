import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook

class Report_Download():
    global writer

    def report_download(self,df, sheetname):
        # Load the Excel workbook
        uploaded =df
        print(uploaded)
        # wb = openpyxl.load_workbook(uploaded, data_only=True)
        sheet_name = sheetname
        XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response = HttpResponse(content_type=XLSX_MIME)
        response['Content-Disposition'] = 'attachment; filename=file_name'
        writer = pd.ExcelWriter(response, engine='xlsxwriter')

        df.to_excel(writer, sheet_name='Report_Download', startrow=1, header=False, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Report_Download']

        # Define a cell format for text with color
        success_format = workbook.add_format({'font_color': '#008000', 'bold': True})
        fail_format = workbook.add_format({'font_color': '#FF0000', 'bold': True})
        message_success_format = workbook.add_format({'bg_color': '#00FF00'})#message success
        # message_fail_format = workbook.add_format({'bg_color': '#FF0000'}) #message fail
        # errorr_success_format=workbook.add_format({'bg_color': '#00FF00'})#error pass
        # error_fail_format=workbook.add_format({'bg_color': '#FF0000'})#error fail
        # for row_num, (status, message,error) in enumerate(zip(df['Status'], df['Message'],df['Error']), start=1):
        for row_num, (status) in enumerate(zip(df['Status'])):
            if status == 'SUCCESS':
                worksheet.write(row_num, df.columns.get_loc('Status'), status, success_format)
                # worksheet.write(row_num, df.columns.get_loc('Message'), message, message_success_format)
                # worksheet.write(row_num, df.columns.get_loc('Error'), error, errorr_success_format)

            elif status == 'Fail':
                worksheet.write(row_num, df.columns.get_loc('Status'), status, fail_format)
                # worksheet.write(row_num, df.columns.get_loc('Message'), message, message_fail_format)
                # worksheet.write(row_num, df.columns.get_loc('Error'), error, error_fail_format)

        header_format = workbook.add_format({
            'bold': True,
            'fg_color': '#D3D3D3',
            'border': 1})
        for col_num, value in enumerate(df.columns):
            worksheet.write(0, col_num, value, header_format)
            # Set column widths based on content length
        for col_num, column in enumerate(df.columns):
            column_len = max(df[column].astype(str).str.len().max(), len(column)) + 3
            worksheet.set_column(col_num, col_num, column_len)
        writer._save()
        return HttpResponse(response)

# if isinstance(value, list):
        # for idx, val in enumerate(value):
        #     if idx > 0 and row_trackers.get(idx, False) is not True:
        #         page.get_by_role("button", name="ADD ROW").click()
        #         page.wait_for_timeout(1000)
        #         row_trackers[idx] = True  # Track that we've added this row
        #
        #     locator = page.locator(f'[formcontrolname="{x}"]').nth(idx)
        #     locator.wait_for(state="attached")
        #
        #     try:
        #         locator.fill(val)
        #         page.wait_for_timeout(500)
        #         locator.click()
        #         options = page.locator('mat-option')
        #         try:
        #             page.wait_for_selector("mat-option",
        #                                    timeout=1000)  # Will timeout if not a dropdown
        #             page.locator(f'mat-option >> text="{val}"').click()
        #         except:
        #             page.locator('mat-option')
        #             page.locator(f'mat-option >> text="{val}"').click()
        #
        #     except Exception:
        #         locator.click()
        #         page.locator(f'mat-option >> text="{val}"').click()







    # def report_download(self,df, sheetname, status_ary, url_ary, Message_ary, error_ary):
    #     # Load the Excel workbook
    #     uploaded_file =df
    #     wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    #     sheet_name = sheetname
    #
    #     # Iterate over the arrays and perform actions
    #     for i in range(len(status_ary)):
    #         status = status_ary[i]
    #         url = url_ary[i]
    #         message = Message_ary[i]
    #         error = error_ary[i]
    #
    #         # Update the Excel workbook (assuming you want to write status and message to it)
    #         ws = wb[sheet_name]
    #         row_to_update = i + 1  # Update the appropriate row in the worksheet
    #         ws.cell(row=row_to_update, column=1, value=status)
    #         ws.cell(row=row_to_update, column=2, value=message)
    #         ws.cell(row=row_to_update, column=3, value=url)
    #         ws.cell(row=row_to_update, column=4, value=error)
    #
    #     # Create a report (in this example, convert data to a DataFrame)
    #     data = {
    #         'Status': status_ary,
    #         'Message': Message_ary,
    #         'URL': url_ary,
    #         'Error': error_ary
    #     }
    #     report_df = pd.DataFrame(data)
    #
    #     # Convert the DataFrame to Excel and add it to the workbook as a new sheet
    #     report_sheet = openpyxl.utils.dataframe.dataframe_to_rows(report_df, index=False)
    #     new_ws = wb.create_sheet(title="Report")
    #     for row in report_sheet:
    #         new_ws.append(row)
    #
    #     # Save the updated workbook to a temporary file
    #     with tempfile.NamedTemporaryFile(delete=False) as temp_file:
    #         wb.save(temp_file.name)
    #
    #     # Serve the temporary file for download
    #     with open(temp_file.name, 'rb') as excel_file:
    #         response = HttpResponse(excel_file.read(),
    #                                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #         response['Content-Disposition'] = 'attachment; filename="your_filename.xlsx"'
    #
    #     return response

    # def report_download(self, request, sheetname, status_ary, url_ary, Message_ary, error_ary):
    #     # Load the Excel workbook
    #     wb = openpyxl.load_workbook(request.FILES['file'])
    #     sheet_name = sheetname
    #
    #     # Iterate over the arrays and perform actions
    #     for i in range(len(status_ary)):
    #         status = status_ary[i]
    #         url = url_ary[i]
    #         message = Message_ary[i]
    #         error = error_ary[i]
    #
    #         # Update the Excel workbook (assuming you want to write status and message to it)
    #         ws = wb[sheet_name]
    #         row_to_update = i + 1  # Update the appropriate row in the worksheet
    #         ws.cell(row=row_to_update, column=1, value=status)
    #         ws.cell(row=row_to_update, column=2, value=message)
    #         ws.cell(row=row_to_update, column=3, value=url)
    #         ws.cell(row=row_to_update, column=4, value=error)
    #
    #     # Create a report (in this example, convert data to a DataFrame)
    #     data = {
    #         'Status': status_ary,
    #         'Message': Message_ary,
    #         'URL': url_ary,
    #         'Error': error_ary
    #     }
    #     report_df = pd.DataFrame(data)
    #
    #     # Convert the DataFrame to Excel and add it to the workbook as a new sheet
    #     report_sheet = openpyxl.utils.dataframe.dataframe_to_rows(report_df, index=False)
    #     new_ws = wb.create_sheet(title="Report")
    #     for row in report_sheet:
    #         new_ws.append(row)
    #
    #     # Save the updated workbook to a virtual file
    #     virtual_workbook = save_virtual_workbook(wb)
    #
    #     # Create an HTTP response to send the file for download
    #     response = HttpResponse(virtual_workbook,
    #                             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #     response['Content-Disposition'] = 'attachment; filename="your_filename.xlsx"'
    #
    #     return response

    # def report_download(self, request, sheetname, status_ary, url_ary, Message_ary,error_ary):
    #     # Load the Excel workbook
    #     wb = openpyxl.load_workbook(request.FILES['file'])
    #     sheet_name = sheetname  # You've already defined this argument
    #
    #     # Iterate over the arrays and perform actions
    #     for i in range(len(status_ary)):
    #         status = status_ary[i]
    #         url = url_ary[i]
    #         message = Message_ary[i]
    #         error=error_ary[i]
    #
    #         # Update the Excel workbook (assuming you want to write status and message to it)
    #         ws = wb[sheet_name]
    #         row_to_update = i + 1  # Update the appropriate row in the worksheet
    #         ws.cell(row=row_to_update, column=1, value=status)
    #         ws.cell(row=row_to_update, column=2, value=message)
    #         ws.cell(row=row_to_update, column=3, value=url)
    #         ws.cell(row=row_to_update, column=4, value=error)
    #
    #         # Create a report (in this example, convert data to a DataFrame)
    #         data = {
    #             'Status': status_ary,
    #             'Message': Message_ary,
    #             'URL': url_ary,
    #             'Error':error_ary
    #
    #         }
    #         report_df = pd.DataFrame(data)
    #
    #         # Convert the DataFrame to Excel and add it to the workbook as a new sheet
    #         report_sheet = openpyxl.utils.dataframe.dataframe_to_rows(report_df, index=False)
    #         new_ws = wb.create_sheet(title="Report")
    #         for row in report_sheet:
    #             new_ws.append(row)
    #
    #     # Save the updated workbook
    #     wb.save()
    #     return new_ws








    # def report_download(self, request, uploaded_file, file_name, sheetname,status,url,Message):
    #     wb = openpyxl.load_workbook(request.FILES['file'])
    #     sheet_names = wb.sheetnames
    #     sheetname=sheetnamex
    #     df = pd.read_excel(uploaded_file, sheet_name=sheetname)
    #
    #     XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #     response = HttpResponse(content_type=XLSX_MIME)
    #     response['Content-Disposition'] = 'attachment; filename=file_name'
    #     writer = pd.ExcelWriter(response, engine='xlsxwriter')
    #     final_df = df
    #     final_df.index.name = 'S.No'
    #     final_df = final_df.reset_index()
    #     final_df['S.No'] = final_df['S.No'] + 1
    #     final_df['Status'] = status
    #     final_df['Remarks'] = "AUTOMATION COMPLETED"
    #     final_df['Screenshot'] = url
    #     final_df['Message']=Message
    #
    #     final_df.to_excel(writer, sheet_name='Report_Download', startrow=1, header=False, index=False)
    #     workbook = writer.book
    #     worksheet = writer.sheets['Report_Download']
    #
    #     # Define a cell format for text with color
    #     success_format = workbook.add_format({'font_color': '#008000'})
    #     fail_format = workbook.add_format({'font_color': '#ff0000'})
    #
    #     # Iterate through the 'Status' column and add formatting to the 'SUCCESS' text
    #     for row_num, value in enumerate(final_df['Status'], start=1):
    #         if value == 'SUCCESS':
    #             worksheet.write(row_num, final_df.columns.get_loc('Status'), value, success_format)
    #         elif value == 'FAIL':
    #             worksheet.write(row_num, final_df.columns.get_loc('Status'), value, fail_format)
    #
    #     for row in worksheet.iter_rows(min_col=worksheet['message'].column, max_col=worksheet['message'].column):
    #         for cell in row:
    #             if "specific_text" in str(cell.value):
    #                 fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    #             else:
    #                 fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    #             cell.fill = fill
    #
    #
    #     header_format = workbook.add_format({
    #         'bold': True,
    #         'fg_color': '#00FF00',
    #         'border': 1})
    #     for col_num, value in enumerate(final_df.columns):
    #         worksheet.write(0, col_num, value, header_format)
    #         # Set column widths based on content length
    #     for col_num, column in enumerate(final_df.columns):
    #         column_len = max(final_df[column].astype(str).str.len().max(), len(column)) + 3
    #         worksheet.set_column(col_num, col_num, column_len)
    #     writer._save()
    #     return HttpResponse(response)

    # def report_download_reject(self, request, sheetname, status_ary, url_ary, Message_ary,error_excpt_ary):
    #     # Load the Excel workbook
    #     wb = openpyxl.load_workbook(request.FILES['file'])
    #     sheet_name = sheetname  # You've already defined this argument
    #
    #     # Iterate over the arrays and perform actions
    #     for i in range(len(status_ary)):
    #         status = status_ary[i]
    #         url = url_ary[i]
    #         message = Message_ary[i]
    #         error=error_excpt_ary[i]
    #
    #         # Update the Excel workbook (assuming you want to write status and message to it)
    #         ws = wb[sheet_name]
    #         row_to_update = i + 1  # Update the appropriate row in the worksheet
    #         ws.cell(row=row_to_update, column=1, value=status)
    #         ws.cell(row=row_to_update, column=2, value=message)
    #         ws.cell(row=row_to_update, column=3, value=url)
    #         ws.cell(row=row_to_update, column=4, value=error)
    #
    #         # Create a report (in this example, convert data to a DataFrame)
    #         data = {
    #             'Status': status_ary,
    #             'Message': Message_ary,
    #             'URL': url_ary,
    #             'Error':error_excpt_ary
    #
    #         }
    #         report_df = pd.DataFrame(data)
    #
    #         # Convert the DataFrame to Excel and add it to the workbook as a new sheet
    #         report_sheet = openpyxl.utils.dataframe.dataframe_to_rows(report_df, index=False)
    #         new_ws = wb.create_sheet(title="Report")
    #         for row in report_sheet:
    #             new_ws.append(row)
    #
    #     # Save the updated workbook
    #     wb.save()
    #     return HttpResponse(response)



    # def report_download_reject(self, request, uploaded_file, file_name, sheetname, status, url,Message,error):
    #     wb = openpyxl.load_workbook(request.FILES['file'])
    #     sheet_names = wb.sheetnames
    #     sheetname = sheetname
    #     df = pd.read_excel(uploaded_file, sheet_name=sheetname)
    #
    #     XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #     response = HttpResponse(content_type=XLSX_MIME)
    #     response['Content-Disposition'] = 'attachment; filename=file_name'
    #     writer = pd.ExcelWriter(response, engine='xlsxwriter')
    #     final_df = df
    #     final_df.index.name = 'S.No'
    #     final_df = final_df.reset_index()
    #     final_df['S.No'] = final_df['S.No'] + 1
    #     final_df['Status'] = status
    #     final_df['Screenshot'] = url
    #     final_df['Message'] = Message
    #     final_df['Remarks'] = error
    #
    #
    #     final_df.to_excel(writer, sheet_name='Report_Download', startrow=1, header=False, index=False)
    #     workbook = writer.book
    #     worksheet = writer.sheets['Report_Download']
    #
    #     # Define a cell format for text with color
    #     success_clr = workbook.add_format({'font_color': '#008000'})
    #     fail_clr = workbook.add_format({'font_color': '#ff0000'})
    #
    #
    #     # Iterate through the 'Status' column and add formatting to the 'SUCCESS' text
    #     for row_num, value in enumerate(final_df['Status'], start=1):
    #         if value == 'SUCCESS':
    #             worksheet.write(row_num, final_df.columns.get_loc('Status'), value, success_clr)
    #         elif value == 'FAIL':
    #             worksheet.write(row_num, final_df.columns.get_loc('Status'), value, fail_clr)
    #     header_format = workbook.add_format({
    #         'bold': True,
    #         'fg_color': '#00FF00',
    #         'border': 1})
    #     for col_num, value in enumerate(final_df.columns):
    #         worksheet.write(0, col_num, value, header_format)
    #         # Set column widths based on content length
    #     for col_num, column in enumerate(final_df.columns):
    #         column_len = max(final_df[column].astype(str).str.len().max(), len(column)) + 3
    #         worksheet.set_column(col_num, col_num, column_len)
    #     writer._save()
    #     return HttpResponse(response)
